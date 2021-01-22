from openpyxl import Workbook # module for creating spreadsheet
import functions

# access work book
wb = Workbook()
# access work sheet
ws = wb.active
filename = "database.xlsx"
# row position
row_pos = 1
ws.title = "User Database"
# worksheet tab color blue
ws.sheet_properties.tabColor = "1072BA"

# headings
ws["A1"] = "Username"
ws["B1"] = "First Name"
ws["C1"] = "Last Name"
ws["D1"] = "Age"
ws["E1"] = "Email"
ws["F1"] = "Password"

while True:

    print()
    new_user = input("Enter New User? (Y/N): ").upper()

    if new_user == "N":
        print()
        print("See you later...")
        break

    elif new_user == "Y":
        # add row
        row_pos += 1

        # cells = row_pos, appropiate column and return value of function
        b = ws.cell(row=row_pos, column=2, value=functions.first_name())
        c = ws.cell(row=row_pos, column=3, value=functions.last_name())
        d = ws.cell(row=row_pos, column=4, value=functions.age_check())
        e = ws.cell(row=row_pos, column=5, value=functions.email_check())
        a = ws.cell(row=row_pos, column=1, value=functions.username_check())
        f = ws.cell(row=row_pos, column=6, value=functions.password_check())

        print()
        print("To add another user to database enter 'y'")
        print()
        print("To save user data to database enter 'n'")
    else:
        print("Invalid, please enter (Y)es or (N)o.")



# saving excel spreadsheet
wb.save(filename)
